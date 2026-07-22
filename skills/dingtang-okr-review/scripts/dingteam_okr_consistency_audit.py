#!/usr/bin/env python3
"""Audit Dingteam scores against a manifest, workbook, and optional online sheet."""

from __future__ import annotations

import argparse
import importlib.util
import json
import subprocess
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


_API_PATH = Path(__file__).resolve().parent / "dingteam_okr_api.py"
_spec = importlib.util.spec_from_file_location("dingteam_okr_api", _API_PATH)
api = importlib.util.module_from_spec(_spec)
assert _spec.loader is not None
_spec.loader.exec_module(api)


def load_manifest(path: str) -> list[dict[str, Any]]:
    payload = json.loads(Path(path).expanduser().read_text(encoding="utf-8"))
    if isinstance(payload, dict):
        payload = payload.get("people")
    if not isinstance(payload, list):
        raise ValueError("manifest must be a list or an object containing people")
    for person in payload:
        if not isinstance(person, dict) or not person.get("person"):
            raise ValueError(f"invalid manifest person: {person!r}")
        if not isinstance(person.get("rows"), list):
            raise ValueError(f"manifest rows missing for {person['person']}")
    return payload


def dws_json(*args: str) -> dict[str, Any]:
    result = subprocess.run(
        ["dws", *args, "--format", "json"],
        text=True,
        capture_output=True,
        check=True,
    )
    payload = json.loads(result.stdout)
    if not isinstance(payload, dict) or not payload.get("success"):
        raise RuntimeError(f"dws read failed: {result.stdout}")
    return payload


def normalized_label(value: object) -> str:
    text = str(value or "").strip()
    for prefix in ("文化：", "文化:", "领导力：", "领导力:"):
        if text.startswith(prefix):
            text = text[len(prefix) :].lstrip()
            break
    return text


def matching_label(value: object, expected_labels: set[str]) -> str:
    text = normalized_label(value)
    matches = [label for label in expected_labels if text.startswith(label)]
    if len(matches) > 1:
        exact = [label for label in matches if text == label]
        if len(exact) == 1:
            return exact[0]
        raise ValueError(f"ambiguous workbook label {text!r}: {matches!r}")
    return matches[0] if matches else ""


def same_value(left: object, right: object, tolerance: float = 0.051) -> bool:
    if str(left or "") == str(right or ""):
        return True
    try:
        return abs(float(left) - float(right)) <= tolerance
    except (TypeError, ValueError):
        return False


def complete_evaluators(kr: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        user
        for role in kr.get("roleUsers", [])
        if isinstance(role, dict)
        for user in role.get("users", [])
        if isinstance(user, dict) and user.get("status") == "complete"
    ]


def sheet_cells(payload: dict[str, Any]) -> list[list[dict[str, Any]]]:
    cells = payload.get("cells")
    if not isinstance(cells, list):
        raise RuntimeError(f"online range response has no cells: {payload!r}")
    return cells


def online_value(cell: object) -> object:
    if isinstance(cell, dict):
        return cell.get("value", "")
    return cell


def audit(
    *,
    okr_id: str,
    manifest_path: str,
    workbook_path: str,
    batch_id: str | None = None,
    online_node: str | None = None,
    label_column: int = 1,
    score_column: int = 6,
    online_start_column: str = "A",
    online_end_column: str = "G",
    allow_browser: bool = True,
) -> dict[str, Any]:
    manifest = load_manifest(manifest_path)
    client = api.ReadClient(allow_browser=allow_browser)
    resolved_batch_id = batch_id or client.resolve_score_batch_id(okr_id)
    workbook = load_workbook(
        Path(workbook_path).expanduser(), read_only=True, data_only=True
    )

    objective_ids = {
        str(row["objectiveId"])
        for person in manifest
        for row in person["rows"]
    }
    system_details = {
        objective_id: client.get_score_detail(objective_id, resolved_batch_id)
        for objective_id in sorted(objective_ids)
    }

    online_ids: dict[str, str] = {}
    if online_node:
        sheet_list = dws_json("sheet", "list", "--node", online_node)
        online_ids = {
            str(sheet["name"]): str(sheet["sheetId"])
            for sheet in sheet_list.get("sheets", [])
            if isinstance(sheet, dict) and sheet.get("name") and sheet.get("sheetId")
        }

    errors: list[dict[str, Any]] = []
    system_kr_count = 0
    workbook_kr_count = 0
    online_cell_count = 0

    def add_error(scope: str, message: str, **context: Any) -> None:
        errors.append({"scope": scope, "message": message, **context})

    for person in manifest:
        name = str(person["person"])
        if name not in workbook.sheetnames:
            add_error("workbook", "person sheet missing", person=name)
            continue
        worksheet = workbook[name]
        expected_labels = {str(row["kr"]).strip() for row in person["rows"]}
        if len(expected_labels) != len(person["rows"]):
            add_error("manifest", "duplicate KR labels for person", person=name)

        local_rows: dict[str, tuple[int, list[object]]] = {}
        for row_index in range(1, worksheet.max_row + 1):
            label = matching_label(
                worksheet.cell(row_index, label_column).value,
                expected_labels,
            )
            if not label:
                continue
            if label in local_rows:
                add_error(
                    "workbook",
                    "duplicate KR label",
                    person=name,
                    kr=label,
                    rows=[local_rows[label][0], row_index],
                )
                continue
            local_rows[label] = (
                row_index,
                [worksheet.cell(row_index, col).value for col in range(1, score_column + 2)],
            )

        expected_evaluators = sorted(str(x) for x in person.get("systemEvaluators", []))
        for expected in person["rows"]:
            objective_id = str(expected["objectiveId"])
            kr_id = str(expected["krId"])
            kr_name = str(expected["kr"])
            objective = system_details.get(objective_id)
            if objective is None:
                add_error(
                    "system", "objective missing", person=name, objectiveId=objective_id
                )
                continue
            kr = next(
                (
                    item
                    for item in objective.get("krScoreDetails", [])
                    if str(item.get("krId")) == kr_id
                ),
                None,
            )
            if kr is None:
                add_error("system", "KR missing", person=name, kr=kr_name, krId=kr_id)
                continue
            evaluators = complete_evaluators(kr)
            actual_names = sorted(str(item.get("name") or item.get("userName") or "") for item in evaluators)
            if not evaluators:
                add_error("system", "no completed evaluator", person=name, kr=kr_name)
            elif expected_evaluators and actual_names != expected_evaluators:
                add_error(
                    "system",
                    "completed evaluator mismatch",
                    person=name,
                    kr=kr_name,
                    expected=expected_evaluators,
                    actual=actual_names,
                )
            else:
                system_kr_count += 1
                evaluator = evaluators[0]
                expected_score = expected.get("score")
                actual_score = evaluator.get("score")
                if actual_score is None or not same_value(
                    float(actual_score) * 100, expected_score
                ):
                    add_error(
                        "system",
                        "score mismatch",
                        person=name,
                        kr=kr_name,
                        expected=expected_score,
                        actual=actual_score,
                    )
                if evaluator.get("info") != expected.get("scoreDesc"):
                    add_error(
                        "system", "comment mismatch", person=name, kr=kr_name
                    )

            local = local_rows.get(kr_name)
            if local is None:
                add_error("workbook", "KR row missing", person=name, kr=kr_name)
            else:
                workbook_kr_count += 1
                actual_local_score = worksheet.cell(local[0], score_column).value
                if not same_value(actual_local_score, expected.get("score")):
                    add_error(
                        "workbook",
                        "score mismatch",
                        person=name,
                        kr=kr_name,
                        expected=expected.get("score"),
                        actual=actual_local_score,
                    )

        if not online_node:
            continue
        if name not in online_ids:
            add_error("online", "person sheet missing", person=name)
            continue
        assessment_rows = sorted(row for row, _ in local_rows.values())
        if not assessment_rows:
            add_error("online", "no assessment rows to compare", person=name)
            continue
        start, end = min(assessment_rows), max(assessment_rows)
        online = sheet_cells(
            dws_json(
                "sheet",
                "range",
                "read",
                "--node",
                online_node,
                "--sheet-id",
                online_ids[name],
                "--range",
                f"{online_start_column}{start}:{online_end_column}{end}",
            )
        )
        for offset, row_index in enumerate(range(start, end + 1)):
            expected_row = [
                worksheet.cell(row_index, col).value
                for col in range(1, score_column + 2)
            ]
            actual_row = online[offset] if offset < len(online) else []
            for col_index, expected_value in enumerate(expected_row):
                actual_value = (
                    online_value(actual_row[col_index])
                    if col_index < len(actual_row)
                    else ""
                )
                online_cell_count += 1
                if not same_value(actual_value, expected_value):
                    add_error(
                        "online",
                        "cell mismatch",
                        person=name,
                        row=row_index,
                        column=col_index + 1,
                        expected=expected_value,
                        actual=actual_value,
                    )

    return {
        "okrId": okr_id,
        "batchId": resolved_batch_id,
        "people": len(manifest),
        "systemKrVerified": system_kr_count,
        "workbookKrVerified": workbook_kr_count,
        "onlineCellsVerified": online_cell_count,
        "errorCount": len(errors),
        "errors": errors,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--okr-id", required=True)
    parser.add_argument("--manifest", required=True)
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--batch-id")
    parser.add_argument("--online-node")
    parser.add_argument("--label-column", type=int, default=1)
    parser.add_argument("--score-column", type=int, default=6)
    parser.add_argument("--online-start-column", default="A")
    parser.add_argument("--online-end-column", default="G")
    parser.add_argument("--output")
    parser.add_argument("--no-browser", action="store_true")
    args = parser.parse_args()

    result = audit(
        okr_id=args.okr_id,
        manifest_path=args.manifest,
        workbook_path=args.workbook,
        batch_id=args.batch_id,
        online_node=args.online_node,
        label_column=args.label_column,
        score_column=args.score_column,
        online_start_column=args.online_start_column,
        online_end_column=args.online_end_column,
        allow_browser=not args.no_browser,
    )
    output = json.dumps(result, ensure_ascii=False, indent=2) + "\n"
    if args.output:
        path = Path(args.output).expanduser().resolve()
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(output, encoding="utf-8")
    print(output, end="")
    return 1 if result["errors"] else 0


if __name__ == "__main__":
    raise SystemExit(main())
